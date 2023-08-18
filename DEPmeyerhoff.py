# Meyerhoff Scholars Program Data Entry Portal
# Authored by Amar McLean; SC04296
import openpyxl


# Press ⌃R to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.

class Excel:
    def __init__(self, file_path):
        self.path = file_path

        # Opens the Excel Sheet
        sheet = openpyxl.load_workbook(file_path)
        self.inactive_sheet = sheet
        self.sheet = sheet.active

        # Find number of columns and rows
        self.columns = self.sheet.max_column
        self.rows = self.sheet.max_row

    def prompt(self):
        user_input = input(
            'What would you like to do?\n\t1. Add to Cohort\n\t2. Edit Cohort\n\t3. Display Cohort\n Input: ')
        if user_input == '1':
            self.add_to_cohort()
        elif user_input == '2':
            self.edit_cohort()
        elif user_input == '3':
            self.display_cohort()

    def add_to_cohort(self):
        return self

    def search_cohort(self):
        cont = True
        variables = {}

        # Find variables to represent each cohort
        print('Which variable would you like to use as an identifier?')
        for x in range(1, self.columns+1):
            var = self.sheet.cell(row=1, column=x).value
            print(f'\t{x}. {var}')
            variables[x] = [var, False, '']

        # Choose the specific variable
        while cont:
            variables[int(input('Variable number: '))][1] = True
            if input('Another(y/n)? ').lower() == 'n':
                cont = False

        # Ask for values of each specified variable
        for x in variables:
            if variables[x][1]:
                variables[x][2] = input(f'What is their {self.sheet.cell(row=1, column=int(x)).value}? ').lower()

        # Find the row number
        possible_row = []
        for y in range(2, self.rows+1):
            wrong = False
            for i in variables:
                if variables[i][1] & (variables[i][0] == self.sheet.cell(row=1, column=i).value) & (
                        variables[i][2] != str(self.sheet.cell(row=y, column=i).value).lower()):
                    wrong = True
                    break
            if not wrong:
                print('found')
                possible_row.append(y)

        return possible_row

    def edit_cohort(self):
        val = self.search_cohort()
        # Change cohort's values
        print('Which variable would you like change?')
        for x in range(1, self.columns + 1):
            var = self.sheet.cell(row=1, column=x).value
            print(f'\t{x}. {var}')

        # Choose the specific variable
        cont = True
        while cont:
            variable = input('Choose number or type \'none\': ').lower()
            if variable == 'none':
                cont = False
            else:
                self.sheet.cell(row=val[0], column=int(variable)).value = input('New value: ')

        self.inactive_sheet.save(self.path)
        self.prompt()

    def display_cohort(self):
        val = self.search_cohort()
        # Display the cohort's values
        for y in val:
            for x in range(1, self.columns+1):
                print(f'{self.sheet.cell(row=1, column=x).value}:{self.sheet.cell(row=y, column=x).value}', end='   ')
            print('\n')

        self.prompt()

    def create_sheet(self):
        print(self)

    def save_sheet(self):
        self.inactive_sheet.save(self.path)

    def __del__(self):
        # self.sheet.close()
        print('DONE')


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    excel_sheet = Excel('big_cheese.xlsx')
    excel_sheet.prompt()
