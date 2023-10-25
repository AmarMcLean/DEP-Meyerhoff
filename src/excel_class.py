# excel_class.py
import openpyxl

class Excel:
    def __init__(self, file_path):
        self.path = file_path

        # Opens the Excel Sheet
        sheet = openpyxl.load_workbook(file_path)
        self.inactive_sheet = sheet
        self.sheet = sheet.active

        # Find number of columns and rows
        self.columns = self.sheet.max_column
        self.rows = self.sheet.max_row

    def add_to_cohort(self):
        return self

    def create_sheet(self):
        print(self)

    def save_sheet(self):
        self.inactive_sheet.save(self.path)

    def __del__(self):
        # self.sheet.close()
        print('DONE')
